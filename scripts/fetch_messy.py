#!/usr/bin/env python3
"""
Fetch a small set of intentionally 'messy' Belize datasets, ZIP them, and write metadata.

- Reads the 'messy' section from catalog.yml
- Each item can be a direct file URL (xlsx/xls/csv) OR a page URL.
  * If it's a page URL, we discover the first .xlsx link on that page.
- Saves raw files under data/messy/raw/<slug>/
- Writes:
    data/messy/_manifest.json         (file-level metadata)
    data/messy/_report.json           (messiness heuristics: merged cells, multi-header, etc.)
    data/messy/_bundle.zip            (all raw files + a README)
    data/messy/_dataset_card.md       (card)
"""

import hashlib
import io
import json
import os
import pathlib
import re
import zipfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
import random
import time

import pandas as pd
import requests
import yaml
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

ROOT = pathlib.Path(__file__).resolve().parents[1]
CATALOG = ROOT / "catalog.yml"
OUT_DIR = ROOT / "data" / "messy"
RAW_DIR = OUT_DIR / "raw"

def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")

def ensure_dir(p: pathlib.Path):
    p.mkdir(parents=True, exist_ok=True)

def sha1(b: bytes) -> str:
    import hashlib as _h
    return _h.sha1(b).hexdigest()

# Robust HTTP
def _make_session():
    s = requests.Session()
    retries = int(os.getenv("CARIBDATA_HTTP_RETRIES", "6"))
    backoff = float(os.getenv("CARIBDATA_HTTP_BACKOFF", "0.8"))
    retry = Retry(
        total=retries,
        connect=retries,
        read=retries,
        status=retries,
        backoff_factor=backoff,
        status_forcelist=[429,500,502,503,504],
        allowed_methods=["GET","HEAD","OPTIONS"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=20)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    s.headers["User-Agent"] = "CaribData/1.0 (+github.com/CaribData)"
    return s

SESSION = _make_session()
HTTP_TIMEOUT = float(os.getenv("CARIBDATA_HTTP_TIMEOUT","90"))

def http_get(url: str, **kwargs) -> requests.Response:
    time.sleep(random.uniform(0.05,0.25))
    r = SESSION.get(url, timeout=kwargs.get("timeout", HTTP_TIMEOUT))
    r.raise_for_status()
    return r

def load_catalog() -> Dict[str, Any]:
    with open(CATALOG, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def is_file_url(u: str) -> bool:
    return bool(re.search(r"\.(xlsx|xls|csv)(\?|$)", u, flags=re.I))

def discover_xlsx_link(page_url: str) -> Optional[str]:
    """Fetch page and return first .xlsx href (absolute)."""
    try:
        r = http_get(page_url, timeout=max(HTTP_TIMEOUT, 60))
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.select("a[href]"):
            href = a.get("href", "")
            if re.search(r"\.xlsx(\?|$)", href, flags=re.I):
                # absolutize
                if href.lower().startswith("http"):
                    return href
                from urllib.parse import urljoin
                return urljoin(page_url, href)
    except Exception:
        return None
    return None

def save_bytes(path: pathlib.Path, content: bytes):
    ensure_dir(path.parent)
    path.write_bytes(content)

def analyze_excel_bytes(b: bytes) -> Dict[str, Any]:
    """Heuristics for 'messiness' on Excel files."""
    info: Dict[str, Any] = {"type": "excel", "sheets": [], "merged_cells": {}, "header_row_guess": {}, "notes": []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
        for ws in wb.worksheets:
            sheet_name = ws.title
            info["sheets"].append(sheet_name)
            # merged cells
            m = len(getattr(ws, "merged_cells", []))
            info["merged_cells"][sheet_name] = m
            # simple header-row guess: find first row with majority non-empty string
            guess = None
            for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
                values = [c.value for c in row]
                non_empty = [v for v in values if v not in (None, "")]
                str_like = [v for v in non_empty if isinstance(v, str)]
                if len(non_empty) and (len(str_like) / max(1, len(non_empty))) >= 0.6:
                    guess = row[0].row
                    break
            info["header_row_guess"][sheet_name] = guess
        if not info["sheets"]:
            info["notes"].append("No visible sheets found.")
    except Exception as e:
        info["notes"].append(f"openpyxl error: {e}")
    return info

def analyze_csv_bytes(b: bytes) -> Dict[str, Any]:
    s = b.decode("utf-8", errors="replace")
    lines = s.splitlines()
    sample = "\n".join(lines[:100])
    import csv, io
    try:
        dialect = csv.Sniffer().sniff(sample)
        delim = dialect.delimiter
    except Exception:
        delim = ","
    # crude row-length variability
    lengths = [len(row.split(delim)) for row in lines[:200] if row.strip()]
    var = len(set(lengths))
    return {"type": "csv", "delimiter": delim, "row_length_variability": var}

def build_readme(items: List[Dict[str, Any]]) -> str:
    lines = [
        "# Belize 'Messy' Data Bundle",
        "",
        "This ZIP contains intentionally messy public datasets to test ingest/clean pipelines.",
        "",
        "## Included",
    ]
    for it in items:
        lines.append(f"- **{it.get('name','(unnamed)')}** — source: {it.get('source','')} — slug: `{it.get('slug')}`")
    lines += [
        "",
        "Each file is provided as-downloaded (“raw”). See `_manifest.json` and `_report.json` for details.",
        ""
    ]
    return "\n".join(lines)

def main():
    cfg = load_catalog()
    messy = cfg.get("messy", {})
    if not messy or not messy.get("enabled", True):
        print("Messy ingest disabled.")
        return

    items: List[Dict[str, Any]] = messy.get("items", [])
    ensure_dir(OUT_DIR); ensure_dir(RAW_DIR)

    manifest: Dict[str, Any] = {"generated_at": now_iso(), "items": []}
    report: Dict[str, Any] = {"generated_at": now_iso(), "files": []}
    errors: List[Dict[str, Any]] = []

    for it in items:
        slug = it["slug"]
        src = it["url"]
        name = it.get("name", slug)
        src_is_file = is_file_url(src)
        resolved = src
        try:
            if not src_is_file:
                # discover .xlsx link on the page
                link = discover_xlsx_link(src)
                if not link:
                    raise RuntimeError("No .xlsx link discovered on page")
                resolved = link

            r = http_get(resolved, timeout=max(HTTP_TIMEOUT, 120))
            b = r.content
            ct = r.headers.get("Content-Type","")
            # infer filename
            from urllib.parse import urlparse, unquote
            fn = pathlib.Path(unquote(urlparse(resolved).path)).name or f"{slug}.bin"
            dest = RAW_DIR / slug / fn
            save_bytes(dest, b)

            # analyze
            lower = fn.lower()
            analysis = {}
            if lower.endswith(".xlsx") or lower.endswith(".xls"):
                analysis = analyze_excel_bytes(b)
            elif lower.endswith(".csv"):
                analysis = analyze_csv_bytes(b)
            else:
                analysis = {"type": "binary/other"}

            manifest["items"].append({
                "slug": slug,
                "name": name,
                "source": it.get("source",""),
                "license": it.get("license","unknown"),
                "original_url": src,
                "resolved_download_url": resolved,
                "saved_path": str(dest.relative_to(ROOT)),
                "size_bytes": len(b),
                "sha1": sha1(b),
                "content_type": ct
            })
            report["files"].append({
                "slug": slug,
                "name": name,
                "analysis": analysis,
                "expected_issues": it.get("expected_issues", [])
            })
        except Exception as e:
            errors.append({"slug": slug, "url": src, "error": str(e)})
            continue

    # write metadata
    (OUT_DIR / "_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (OUT_DIR / "_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")
    if errors:
        (OUT_DIR / "_errors.json").write_text(json.dumps(errors, indent=2), encoding="utf-8")

    # dataset card (one-time)
    card = OUT_DIR / "_dataset_card.md"
    if not card.exists():
        card.write_text(
            "# Belize 'Messy' Datasets (Test Bundle)\n\n"
            "Raw Excel/CSV files pulled from public sources to exercise ingest and cleaning.\n\n"
            "See `_manifest.json` for file list and `_report.json` for structural hints (sheets, merged cells, etc.).\n",
            encoding="utf-8"
        )

    # bundle ZIP
    bundle = OUT_DIR / "_bundle.zip"
    with zipfile.ZipFile(bundle, "w", zipfile.ZIP_DEFLATED) as z:
        # add a README listing
        readme = build_readme(items)
        z.writestr("README.md", readme)
        # add metadata
        z.writestr("_manifest.json", json.dumps(manifest, indent=2))
        z.writestr("_report.json", json.dumps(report, indent=2))
        # add raw files
        for p in RAW_DIR.rglob("*"):
            if p.is_file():
                z.write(p, p.relative_to(OUT_DIR).as_posix())

    print("Messy bundle written:", bundle)

if __name__ == "__main__":
    main()
